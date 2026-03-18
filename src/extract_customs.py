"""
extract_customs.py
------------------
Pure Claude vision extraction — no pdfplumber, no positional heuristics.
Renders PDF pages as images and sends them directly to Claude.

Requirements:
    pip install pymupdf anthropic openpyxl
"""

import sys, re, json, base64, os
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── 1. PDF → IMAGES ──────────────────────────────────────────────────────────

def pdf_to_images(pdf_path: str, page_indices: list, dpi: int = 150) -> list:
    import pymupdf
    doc = pymupdf.open(pdf_path)
    mat = pymupdf.Matrix(dpi / 72, dpi / 72)
    imgs = []
    for i in page_indices:
        if i < len(doc):
            pix = doc[i].get_pixmap(matrix=mat)
            imgs.append(pix.tobytes("png"))
    doc.close()
    return imgs


def find_part2_end(pdf_path: str) -> int:
    """Return index of first page containing PART - III (= exclusive end of Part II)."""
    import pymupdf
    doc = pymupdf.open(pdf_path)
    result = len(doc)
    for i, page in enumerate(doc):
        if i > 0 and "PART - III" in page.get_text():
            result = i
            break
    doc.close()
    return result


def num_pages(pdf_path: str) -> int:
    import pymupdf
    doc = pymupdf.open(pdf_path)
    n = len(doc)
    doc.close()
    return n


# ── 2. API ────────────────────────────────────────────────────────────────────

def get_api_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    key_file = os.path.join(
        os.path.dirname(sys.executable) if getattr(sys, "frozen", False)
        else os.path.dirname(os.path.abspath(__file__)),
        "key.txt"
    )
    if os.path.exists(key_file):
        with open(key_file) as f:
            return f.read().strip()
    import tkinter as tk
    from tkinter import simpledialog, messagebox
    root = tk.Tk(); root.withdraw()
    key = simpledialog.askstring(
        "API Key Required",
        "Enter your Anthropic API key.\n\nTip: put it in key.txt next to this app.",
        show="*"
    )
    if not key:
        messagebox.showerror("Error", "No API key provided. Exiting.")
        sys.exit(1)
    if messagebox.askyesno("Save key?", "Save to key.txt so you're not asked again?"):
        with open(key_file, "w") as f:
            f.write(key)
    return key


def call_claude(prompt: str, images: list, max_tokens: int = 2048) -> str:
    client  = anthropic.Anthropic(api_key=get_api_key())
    content = []
    for img in images:
        content.append({"type": "image", "source": {
            "type": "base64", "media_type": "image/png",
            "data": base64.standard_b64encode(img).decode()
        }})
    content.append({"type": "text", "text": prompt})
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=max_tokens,
        system=(
            "You are a data extraction assistant. "
            "Always respond with valid JSON only — no markdown, no explanation. "
            "In string values never use raw double-quote characters; "
            "replace inch marks with 'in' and & with 'and'."
        ),
        messages=[{"role": "user", "content": content}]
    )
    return resp.content[0].text.strip()


def parse_json(raw: str):
    raw = re.sub(r"^```[a-z]*\n?", "", raw.strip()).rstrip("`").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Try extracting outermost { } or [ ]
        for s, e in [('{', '}'), ('[', ']')]:
            i, j = raw.find(s), raw.rfind(e)
            if i != -1 and j > i:
                try:
                    return json.loads(raw[i:j+1])
                except json.JSONDecodeError:
                    pass
        raise


# ── 3. PROMPTS ────────────────────────────────────────────────────────────────

PART1_PROMPT = """You are reading page 1 of an Indian Customs EDI System shipping bill.

Extract EXACTLY these fields and return a single JSON object.

Field locations:
- port_code, sb_no, sb_date: top-right header box
- fob_value, freight, insurance, discount, com: C.VALU SUMMA row 1 values (numbers below the headers, left to right)
- deductions, pc, duty, cess: C.VALU SUMMA row 2 values (numbers below the headers, left to right)
- dbk_claim, igst_amt, cess_amt: D.EX.PR row 1 values
- igst_value, rodtep_amt, rosctl_amt: D.EX.PR row 2 values
- inv_no, inv_amt, currency: F.INVOICE SUMMARY section

IMPORTANT: All numeric fields must contain only the number — never a header label like "7.P/C" or "8.DUTY".
If a value cell appears empty, use "0".

Return ONLY this JSON (no markdown):
{
  "port_code": "",
  "sb_no": "",
  "sb_date": "",
  "fob_value": "",
  "freight": "",
  "insurance": "",
  "discount": "",
  "com": "",
  "deductions": "",
  "pc": "",
  "duty": "",
  "cess": "",
  "dbk_claim": "",
  "igst_amt": "",
  "cess_amt": "",
  "igst_value": "",
  "rodtep_amt": "",
  "rosctl_amt": "",
  "inv_no": "",
  "inv_amt": "",
  "currency": ""
}"""

PART2_VALDTLS_PROMPT = """You are reading a Part II invoice page of an Indian Customs EDI System shipping bill.

Extract ONLY the C.VAL DTLS row (near the top, labeled C.VAL DTLS).
Column order strictly left to right:
Invoice Value | FOB Value | Freight | Insurance | Discount | Commission | Deduct | P/C | Exchange Rate

Currency labels (e.g. USD) appear below Invoice Value and FOB Value.
Exchange Rate looks like "1 USD INR 81.4".
Values must be numbers only — never output header text as a value.

Return ONLY this JSON:
{
  "invoice_value": "",
  "invoice_currency": "",
  "fob_value": "",
  "fob_currency": "",
  "freight": "",
  "insurance": "",
  "discount": "",
  "commission": "",
  "deduct": "",
  "pc": "",
  "exchange_rate": ""
}"""

PART2_ITEMS_PROMPT = """You are reading Part II invoice detail pages of an Indian Customs EDI System shipping bill.

Extract ALL items from the D.ITEM DETAILS table across all pages shown.
Each item row has: Item No, HS Code, Description, Quantity, Rate, Value(F/C).

Rules:
- Combine wrapped/split description lines into one clean string per item
- Replace any inch-mark characters with 'in' (e.g. 2-1/2" becomes 2-1/2in)
- Replace & with 'and'
- quantity, rate, value_fc must be numbers (not strings)
- Include EVERY item — do not stop early

Return ONLY a JSON array:
[
  {
    "item_no": 1,
    "hs_code": "73089090",
    "description": "clean description",
    "quantity": 750,
    "rate": 1.546,
    "value_fc": 1159.5
  }
]"""


# ── 4. MAIN EXTRACTION ────────────────────────────────────────────────────────

def extract_from_pdf(pdf_path: str) -> tuple:
    """Returns (summary, val_dtls, items). Pure Claude vision — no pdfplumber."""
    total = num_pages(pdf_path)
    part2_end = find_part2_end(pdf_path)   # first page index of PART III

    # ── Part I: page 0 only ───────────────────────────────────────────────────
    p1_imgs = pdf_to_images(pdf_path, [0])
    summary = {}
    for attempt in range(2):
        try:
            summary = parse_json(call_claude(PART1_PROMPT, p1_imgs, max_tokens=1024))
            break
        except (ValueError, json.JSONDecodeError):
            pass

    # ── Part II val_dtls: page 1 only ────────────────────────────────────────
    val_dtls = {}
    if total > 1:
        p2_img = pdf_to_images(pdf_path, [1])
        for attempt in range(2):
            try:
                val_dtls = parse_json(call_claude(PART2_VALDTLS_PROMPT, p2_img, max_tokens=512))
                break
            except (ValueError, json.JSONDecodeError):
                pass

    # ── Part II items: pages 1..part2_end-1 (cap at 4 pages) ─────────────────
    items = []
    if total > 1:
        item_idxs = list(range(1, min(part2_end, 5)))
        item_imgs = pdf_to_images(pdf_path, item_idxs)
        for attempt in range(2):
            try:
                result = parse_json(call_claude(PART2_ITEMS_PROMPT, item_imgs, max_tokens=8192))
                items  = result if isinstance(result, list) else result.get("items", [])
                break
            except (ValueError, json.JSONDecodeError):
                pass

    return summary, val_dtls, items


# ── 5. EXCEL OUTPUT ───────────────────────────────────────────────────────────

def write_to_sheet(ws, summary: dict, val_dtls: dict, items: list, start_row: int) -> int:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def cell_style(cell, value, bold=False, bg=None, font_color="000000", align=None, nfmt=None):
        cell.value = value
        cell.font  = Font(name="Arial", bold=bold, size=10, color=font_color)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if align:
            cell.alignment = align
        if nfmt:
            cell.number_format = nfmt

    # Header row
    hr = start_row
    ws.row_dimensions[hr].height = 13.9
    header_cols = [
        (1,  "Port Code",        "DEEAF1"),
        (2,  "SB No",            None),
        (3,  "SB Date",          "DEEAF1"),
        (6,  "FOB Value",        "DEEAF1"),
        (7,  "Freight",          None),
        (8,  "Insurance",        "DEEAF1"),
        (9,  "Discount",         None),
        (10, "COM",              "DEEAF1"),
        (11, "Deductions",       None),
        (12, "P/C",              "DEEAF1"),
        (13, "Duty",             None),
        (14, "Cess",             "DEEAF1"),
        (17, "DBK Claim",        "DEEAF1"),
        (18, "IGST AMT",         None),
        (19, "Cess AMT",         "DEEAF1"),
        (20, "IGST Value",       None),
        (21, "RODTEP AMT",       "DEEAF1"),
        (22, "ROSCTL AMT",       None),
        (25, "INV No",           "DEEAF1"),
        (26, "INV AMT",          None),
        (27, "Currency",         "DEEAF1"),
        (30, "Invoice Value",    "DEEAF1"),
        (31, "Invoice Currency", None),
        (32, "FOB Value",        "DEEAF1"),
        (33, "FOB Currency",     None),
        (34, "Freight",          "DEEAF1"),
        (35, "Insurance",        None),
        (36, "Discount",         "DEEAF1"),
        (37, "Commission",       None),
        (38, "Deduct",           "DEEAF1"),
        (39, "P/C",              None),
        (40, "Exchange Rate",    "DEEAF1"),
    ]
    for col_idx, label, bg in header_cols:
        cell_style(ws.cell(row=hr, column=col_idx), label, bold=True, bg=bg)

    # Value row
    vr = start_row + 1
    value_cols = [
        (1,  summary.get("port_code")),
        (2,  summary.get("sb_no")),
        (3,  summary.get("sb_date")),
        (6,  summary.get("fob_value")),
        (7,  summary.get("freight")),
        (8,  summary.get("insurance")),
        (9,  summary.get("discount")),
        (10, summary.get("com")),
        (11, summary.get("deductions")),
        (12, summary.get("pc")),
        (13, summary.get("duty")),
        (14, summary.get("cess")),
        (17, summary.get("dbk_claim")),
        (18, summary.get("igst_amt")),
        (19, summary.get("cess_amt")),
        (20, summary.get("igst_value")),
        (21, summary.get("rodtep_amt")),
        (22, summary.get("rosctl_amt")),
        (25, summary.get("inv_no")),
        (26, summary.get("inv_amt")),
        (27, summary.get("currency")),
        (30, val_dtls.get("invoice_value")),
        (31, val_dtls.get("invoice_currency")),
        (32, val_dtls.get("fob_value")),
        (33, val_dtls.get("fob_currency")),
        (34, val_dtls.get("freight")),
        (35, val_dtls.get("insurance")),
        (36, val_dtls.get("discount")),
        (37, val_dtls.get("commission")),
        (38, val_dtls.get("deduct")),
        (39, val_dtls.get("pc")),
        (40, val_dtls.get("exchange_rate")),
    ]
    for col_idx, value in value_cols:
        cell_style(ws.cell(row=vr, column=col_idx), value, bg="FFFF00")

    # Item headers
    ihr = start_row + 2
    ws.row_dimensions[ihr].height = 26.25
    cell_style(ws.cell(row=ihr, column=42), "SB No", bold=True)
    for i, hdr in enumerate(["HS Code", "Description", "Quantity", "Rate (USD)", "Value F/C (USD)"]):
        cell_style(ws.cell(row=ihr, column=43 + i), hdr,
                   bold=True, bg="2E75B6", font_color="FFFFFF", align=center)

    # Item rows
    sb_no = summary.get("sb_no", "")
    item_nfmts  = [None, None, "#,##0", "#,##0.000", "#,##0.00"]
    item_aligns = [center, left, center, center, center]
    for idx, item in enumerate(items):
        r       = start_row + 3 + idx
        fill_bg = "DEEAF1" if idx % 2 == 0 else None
        cell_style(ws.cell(row=r, column=42), sb_no, bg="FFFF00")
        values = [item.get("hs_code"), item.get("description"),
                  item.get("quantity"), item.get("rate"), item.get("value_fc")]
        for i, (val, aln, nfmt) in enumerate(zip(values, item_aligns, item_nfmts)):
            cell_style(ws.cell(row=r, column=43 + i), val, bg=fill_bg, align=aln, nfmt=nfmt)

    return start_row + 3 + len(items) + 1


def build_workbook(all_results: list, skipped: list, failures: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customs Extract"

    col_widths = {
        "A": 8.9375, "B": 7.25,   "C": 9.25,
        "F": 9.625,  "G": 6.5625, "H": 8.8125, "I": 8.0625,
        "J": 4.5625, "K": 10.0625,"L": 3.5625, "M": 4.5,   "N": 4.75,
        "Q": 9.75,   "R": 8.875,  "S": 9.8125, "T": 9.8125,
        "U": 12.0625,"V": 12.0,   "Y": 8.5625, "Z": 7.6875,"AA": 8.25,
        "AD": 11.6875,"AE": 14.75,"AF": 9.375, "AG": 12.4375,
        "AH": 6.5625,"AI": 8.8125,"AJ": 8.0625,"AK": 11.25,
        "AL": 6.5,   "AM": 3.5625,"AN": 13.0,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    current_row = 1
    for result in all_results:
        current_row = write_to_sheet(ws, result["summary"], result["val_dtls"],
                                     result["items"], start_row=current_row)

    # Errors sheet
    we = wb.create_sheet("Errors")
    we.column_dimensions["A"].width = 40
    we.column_dimensions["B"].width = 18
    we.column_dimensions["C"].width = 50
    hdr_fill = PatternFill("solid", start_color="C00000")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    row_fill = PatternFill("solid", start_color="FFE0E0")
    nrm_font = Font(name="Arial", size=10)
    ctr      = Alignment(horizontal="center", vertical="center")
    lft      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    for col, label in enumerate(["File", "Type", "Reason"], 1):
        c = we.cell(row=1, column=col, value=label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr
    we.row_dimensions[1].height = 18
    row = 2
    for filename, reason in skipped:
        for col, val in enumerate([filename, "Duplicate", reason], 1):
            c = we.cell(row=row, column=col, value=val)
            c.font = nrm_font; c.fill = row_fill; c.alignment = lft
        we.row_dimensions[row].height = 16; row += 1
    for filename, reason in failures:
        for col, val in enumerate([filename, "Error", reason], 1):
            c = we.cell(row=row, column=col, value=val)
            c.font = nrm_font; c.alignment = lft
        we.row_dimensions[row].height = 16; row += 1

    wb.save(output_path)
    print(f"✅  Saved: {output_path}")


# ── 6. MAIN ───────────────────────────────────────────────────────────────────

def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from datetime import datetime

    root = tk.Tk(); root.withdraw()

    pdf_paths = filedialog.askopenfilenames(
        title="Select Customs PDF files to process",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
    )
    if not pdf_paths:
        messagebox.showinfo("Cancelled", "No files selected."); sys.exit(0)

    output_dir = filedialog.askdirectory(title="Select folder to save Excel output")
    if not output_dir:
        messagebox.showinfo("Cancelled", "No output folder selected."); sys.exit(0)

    progress_win = tk.Tk()
    progress_win.title("Customs PDF Extractor")
    progress_win.geometry("500x300")
    progress_win.resizable(False, False)

    tk.Label(progress_win, text="Processing PDFs...", font=("Arial", 12, "bold")).pack(pady=10)
    progress_bar = ttk.Progressbar(progress_win, length=460, mode="determinate",
                                   maximum=len(pdf_paths))
    progress_bar.pack(pady=5, padx=20)
    status_label = tk.Label(progress_win, text="", font=("Arial", 9), wraplength=460)
    status_label.pack(pady=5)
    log_box = tk.Text(progress_win, height=10, font=("Courier", 8), state="disabled")
    log_box.pack(pady=5, padx=20, fill="both", expand=True)

    def log(msg):
        log_box.config(state="normal")
        log_box.insert("end", msg + "\n")
        log_box.see("end")
        log_box.config(state="disabled")
        progress_win.update()

    all_results, failures, skipped = [], [], []
    seen_sb_nos, seen_inv_nos = set(), set()

    for i, pdf_path in enumerate(pdf_paths):
        filename = os.path.basename(pdf_path)
        status_label.config(text=f"Processing: {filename}")
        log(f"\n[{i+1}/{len(pdf_paths)}] {filename}")
        progress_win.update()

        try:
            log("  → Extracting Part I (Claude vision)...")
            log("  → Extracting Part II val_dtls + items...")
            summary, val_dtls, items = extract_from_pdf(pdf_path)

            sb_no  = str(summary.get("sb_no")  or "").strip()
            inv_no = str(summary.get("inv_no") or "").strip()

            dup_reason = None
            if sb_no  and sb_no  in seen_sb_nos:  dup_reason = f"duplicate SB No ({sb_no})"
            elif inv_no and inv_no in seen_inv_nos: dup_reason = f"duplicate Invoice No ({inv_no})"

            if dup_reason:
                log(f"  ⚠️  SKIPPED — {dup_reason}")
                skipped.append((filename, dup_reason))
                progress_bar["value"] = i + 1; progress_win.update(); continue

            if sb_no:  seen_sb_nos.add(sb_no)
            if inv_no: seen_inv_nos.add(inv_no)

            all_results.append({"summary": summary, "val_dtls": val_dtls, "items": items})
            log(f"  ✅ Extracted {len(items)} items")

        except Exception as e:
            log(f"  ❌ ERROR: {e}")
            failures.append((filename, str(e)))

        progress_bar["value"] = i + 1
        progress_win.update()

    if all_results or skipped or failures:
        status_label.config(text="Building Excel..."); progress_win.update()
        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"customs_extracted_{timestamp}.xlsx")
        log(f"\n→ Writing {len(all_results)} PDF(s) ...")
        try:
            build_workbook(all_results, skipped, failures, output_path)
            log(f"✅ Saved: {output_path}")
        except Exception as e:
            log(f"❌ Failed to save Excel: {e}")

    status_label.config(text="Done!")
    summary_msg = (
        f"✅ {len(all_results)} PDF(s) extracted\n"
        f"⚠️  {len(skipped)} skipped (duplicates)\n"
        f"❌ {len(failures)} failed\n\nSaved to:\n{output_dir}"
    )
    if skipped or failures:
        summary_msg += "\n\nSee the 'Errors' sheet for details."
    log("\n" + "=" * 50); log(summary_msg)
    messagebox.showinfo("Complete", summary_msg)
    progress_win.protocol("WM_DELETE_WINDOW", lambda: (progress_win.destroy(), sys.exit(0)))
    progress_win.mainloop()
    sys.exit(0)


if __name__ == "__main__":
    main()